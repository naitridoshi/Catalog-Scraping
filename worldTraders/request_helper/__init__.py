import json
import multiprocessing
import os
import re
import time
import pandas as pd
import pytz
import requests
import urllib3
from bs4 import BeautifulSoup
from openpyxl.reader.excel import load_workbook

# from common.config import DATA_CENTER_PROXIES
from common.constants import BASIC_HEADERS, BATCH_SIZE, MAX_PROCESSES
from common.custom_logger import color_string, get_logger

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logger, listener = get_logger("WorldTradersRequestHelper")
listener.start()

dubai_tz = pytz.timezone("Asia/Dubai")

class RequestHelper:
    def __init__(self, proxies: dict = None, headers: dict = BASIC_HEADERS):
        self.proxies = proxies
        self.headers = headers
        self.shared_list = multiprocessing.Manager().list()

    def request(self, url: str, method: str = 'GET', timeout: int = 40,params:dict=None,json:dict | list =None,headers:dict=None,data:str|dict|None=None):
        logger.debug(f"Requesting {url} ...")
        for try_request in range(1, 5):
            start_time = time.time()
            try:
                response = requests.request(
                    method=method,
                    url=url,
                    params=params,
                    json=json,
                    data=data,
                    headers=self.headers if headers is None else headers,
                    proxies=self.proxies,
                    verify=False,
                    timeout=timeout,
                    stream=True
                )
                time_taken = f'{time.time() - start_time:.2f} seconds'
                if response.status_code == 200:
                    logger.debug(
                        f'Try: {try_request}, '
                        f'Status Code: {response.status_code}, '
                        f'Response Length: '
                        f'{len(response.text) / 1024 / 1024:.2f} MB, '
                        f'Time Taken: {color_string(time_taken)}.'
                    )
                    return response
                else:
                    logger.warning(
                        f'REQUEST FAILED - {try_request}: '
                        f'Status Code: {response.status_code}, '
                        f'Text: {response.text} '
                        f'Time Taken: {color_string(time_taken)}.'
                    )
            except Exception as err:
                logger.error(
                    f'ERROR OCCURRED - {try_request}: Time Taken '
                    f"{color_string(f'{time.time() - start_time:.2f} seconds')}"
                    f', Error: {err}'
                )

    @staticmethod
    def parse_ipc_data(soup: BeautifulSoup):
        parsed_data = []
        try:
            logger.debug(f"Found {len(soup)} rows to parse ...")
            for count, block in enumerate(soup):
                try:
                    company_name = block.find("h5").get_text(strip=True)

                    contact_person_tag = block.find("h6", string="Contact Person")
                    contact_person = contact_person_tag.find_next_sibling("p").get_text(strip=True) if contact_person_tag else ""

                    email_tag = block.find("a", href=lambda x: x and x.startswith("mailto:"))
                    email = email_tag.get_text(strip=True) if email_tag else ""

                    phone_tag = block.find("a", href=lambda x: x and x.startswith("tel:"))
                    phone = phone_tag.get_text(strip=True) if phone_tag else ""

                    calendar_tag = block.find("em", class_="fa-calendar")
                    date = calendar_tag.find_parent("p").get_text(strip=True).replace("\n","").strip() if calendar_tag else ""

                    seen_by_tag = block.find("em", class_="fa-eye")
                    seen_by = seen_by_tag.find_parent("p").get_text(strip=True).replace("\n","").strip() if seen_by_tag else ""

                    record = [company_name, contact_person, email, phone, date, seen_by]
                    parsed_data.append(record)

                    logger.debug(f"count: {count} - extracted {record}")

                except Exception as inner_e:
                    logger.warning(f"Failed to parse block {count}: {str(inner_e)}")

            logger.info(f"Successfully parsed {len(parsed_data)} records from the table.")

        except Exception as e:
            logger.error(f"Error occurred while parsing IPC data: {str(e)}")

        return parsed_data

    @staticmethod
    def save_to_excel(data_list, filename, url, max_rows_per_sheet=100000):
        try:
            if data_list is None:
                logger.warning("No data to save. The list is empty.")
                raise Exception("EMPTY DATA")

            directory = os.path.dirname(filename)
            if directory and not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
                logger.info(f"Created directory: {directory}")

            headers = ["COMPANY NAME", "CONTACT PERSON", "EMAIL", "PHONE-NUMBER","LAST UPDATED", "SEEN BY"]
            data_list_padded = [row + [""] * (len(headers) - len(row)) for row in data_list]
            new_df = pd.DataFrame(data_list_padded, columns=headers).fillna("")

            if os.path.exists(filename):
                # If file exists, open and check last sheet
                book = load_workbook(filename)
                last_sheet = book.sheetnames[-1]
                existing_df = pd.read_excel(filename, sheet_name=last_sheet)

                if len(existing_df) + len(new_df) <= max_rows_per_sheet:
                    combined_df = pd.concat([existing_df, new_df], ignore_index=True)
                    with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        combined_df.to_excel(writer, index=False, sheet_name=last_sheet)
                    logger.info(f"Appended data to {last_sheet} in {filename}")
                else:
                    # Create a new sheet
                    new_sheet_name = f"{last_sheet}_Part2" if "_Part" not in last_sheet else f"{last_sheet[:-1]}{int(last_sheet[-1]) + 1}"
                    with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
                        new_df.to_excel(writer, index=False, sheet_name=new_sheet_name)
                    logger.info(f"Data exceeded limit, written to new sheet {new_sheet_name} in {filename}")
            else:
                # If file does not exist, create new
                sheet_name = f"IPC-{str(url.split('/')[-2]).upper()[:25]}"
                with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
                    new_df.to_excel(writer, index=False, sheet_name=sheet_name)
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]

                    # Header Format
                    header_format = workbook.add_format({
                        'bold': True,
                        'text_wrap': True,
                        'valign': 'center',
                        'fg_color': '#4F81BD',  # background color
                        'font_color': 'white',  # text color
                        'border': 1
                    })

                    for col_num, header in enumerate(headers):
                        worksheet.write(0, col_num, header, header_format)

                    # Cell Format
                    cell_format = workbook.add_format({
                        'border': 1,
                        'text_wrap': True,
                        'valign': 'top'
                    })

                    # Write Data with Formatting
                    for row_num in range(1, len(new_df) + 1):
                        for col_num in range(len(new_df.columns)):
                            value = new_df.iloc[row_num - 1, col_num]
                            safe_value = "" if pd.isna(value) or str(value).lower() in ["nan", "inf", "-inf"] else value
                            worksheet.write(row_num, col_num, safe_value, cell_format)

                    # Auto Adjust Columns
                    for i, column in enumerate(new_df.columns):
                        max_len = max(new_df[column].astype(str).map(len).max(), len(column))
                        worksheet.set_column(i, i, min(max_len + 2, 30))

                logger.info(f"Data saved and formatted to new workbook {filename}")

        except Exception as e:
            logger.error(f"Error occurred while saving data to Excel: {str(e)}")

    def get_data_from_url_using_soup(self, url: str):
        response = self.request(url)
        if response is None:
            return None, None
        logger.debug(
            f'Got the response for {url}, data length: {len(response.text)}'
        )
        soup = BeautifulSoup(response.text, 'html.parser')
        excel_class=soup.find_all('div',class_="company-info")
        if excel_class:
            logger.info("Sending excel class soup to parse")
            return self.parse_ipc_data(excel_class)
        else:
            logger.warning("Sending original soup to parse")
            return self.parse_ipc_data(soup)

    def main(self, main_url,filename):
        raw_data = self.get_data_from_url_using_soup(main_url)
        self.save_to_excel(raw_data,filename,main_url)

    @staticmethod
    def clean_text_from_json(filename: str):
        try:
            with open(filename, 'r') as f:
                data = json.load(f)
            for item in data:
                item['data'] = re.sub(r'\s+', ' ', item['data'].strip())
            with open('data2.json', 'w') as f:
                json.dump(data, f, indent=4)
            logger.debug('Cleaned text saved to data2.json')
        except (FileNotFoundError, json.JSONDecodeError) as e:
            logger.error(f"Error processing file {filename}: {e}")


if __name__ == '__main__':
    scraper = RequestHelper(
        # proxies=DATA_CENTER_PROXIES,
        headers={}
    )
    _url = 'https://www.cp.pt/passageiros/en/how-to-travel/Useful-information'
    res = scraper.request('https://www.ifconfig.me/all.json')
    print(res.json())
    print(res.status_code)
